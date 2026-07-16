"""标准场景库 API（场景库中心）。

2026-07-13 · Harness P3/P4 底座。挂 /api/scenes:
- GET /api/scenes                 列出场景(可 domain / q 过滤)
- GET /api/scenes/domains         各域场景数概览
- GET /api/scenes/{id}            单场景详情
- GET /api/scenes/{id}/changes    单场景变更历史
- GET /api/scene-changes          全库最近变更历史(何时/哪个项目/新增或优化)

seed_scenes_if_empty:首启从 backend/data/scenes_seed.json 导入标准场景(空表才导)。
"""
import io
import os
import json
import urllib.parse

import structlog
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import Response
from pydantic import BaseModel
from datetime import datetime
from sqlalchemy import select, func

from models import async_session_maker, get_session
from models.scene import StandardScene, SceneChange, AiCapability
from services.auth import get_current_user, require_admin
from models.user import User
from sqlalchemy.ext.asyncio import AsyncSession

logger = structlog.get_logger()
router = APIRouter()

_SEED_PATH = os.path.join(os.path.dirname(__file__), "..", "seeds", "scenes_seed.json")
_AI_SEED_PATH = os.path.join(os.path.dirname(__file__), "..", "seeds", "ai_capabilities_seed.json")


async def seed_ai_capabilities_if_empty() -> None:
    """首启导入纷享 AI 能力目录(仅当 ai_capabilities 为空)。"""
    async with async_session_maker() as s:
        n = (await s.execute(select(func.count(AiCapability.id)))).scalar_one()
        if n > 0:
            return
        if not os.path.exists(_AI_SEED_PATH):
            logger.warning("ai_capabilities_seed_missing", path=_AI_SEED_PATH)
            return
        data = json.load(open(_AI_SEED_PATH, encoding="utf-8"))
        for r in data:
            s.add(AiCapability(
                domain=r.get("domain", ""), agent=r.get("agent", ""), skill=r.get("skill", ""),
                status=r.get("status", ""), plan_date=r.get("plan_date") or None,
                description=r.get("description"), outputs=r.get("outputs") or [],
                sort=r.get("sort", 0),
            ))
        await s.commit()
        logger.info("ai_capabilities_seeded", count=len(data))


async def seed_scenes_if_empty() -> None:
    """首启导入标准场景库(仅当 standard_scenes 为空)。"""
    async with async_session_maker() as s:
        n = (await s.execute(select(func.count(StandardScene.id)))).scalar_one()
        if n > 0:
            return
        if not os.path.exists(_SEED_PATH):
            logger.warning("scenes_seed_missing", path=_SEED_PATH)
            return
        data = json.load(open(_SEED_PATH, encoding="utf-8"))
        for r in data:
            s.add(StandardScene(
                domain=r.get("domain", ""), stage=r.get("stage", ""),
                stage_label=r.get("stage_label"), code=r.get("code", ""),
                name=r.get("name", ""), summary=r.get("stage_def") or r.get("summary"),
                source_type="standard",
            ))
        await s.commit()
        logger.info("scenes_seeded", count=len(data))


class SceneDto(BaseModel):
    id: int
    domain: str
    stage: str
    stage_label: str | None = None
    code: str
    name: str
    summary: str | None = None
    description: str | None = None
    business_rules: str | None = None
    process: str | None = None
    recommended_fields: list = []
    research_questions: list = []    # 关键调研问题(字符串列表)
    tags: list = []
    ai_capabilities: list = []       # 匹配的 AI 能力 id 列表
    source_type: str
    source_project_name: str | None = None
    status: str
    version: int
    updated_at: datetime


class SceneChangeDto(BaseModel):
    id: int
    scene_id: int | None = None
    scene_code: str
    domain: str | None = None
    change_type: str
    project_name: str | None = None
    summary: str | None = None
    created_by: str | None = None
    created_at: datetime


def _scene_dto(x: StandardScene) -> SceneDto:
    return SceneDto(
        id=x.id, domain=x.domain, stage=x.stage, stage_label=x.stage_label,
        code=x.code, name=x.name, summary=x.summary,
        description=x.description, business_rules=x.business_rules, process=x.process,
        recommended_fields=x.recommended_fields or [],
        research_questions=x.research_questions or [], tags=x.tags or [],
        ai_capabilities=x.ai_capabilities or [],
        source_type=x.source_type,
        source_project_name=x.source_project_name, status=x.status, version=x.version,
        updated_at=x.updated_at,
    )


@router.get("/scenes/domains", dependencies=[Depends(get_current_user)])
async def scene_domains(session: AsyncSession = Depends(get_session)):
    """各域场景数(概览卡)。"""
    rows = (await session.execute(
        select(StandardScene.domain, func.count(StandardScene.id))
        .where(StandardScene.status == "active")
        .group_by(StandardScene.domain)
    )).all()
    order = ["LTC", "MTL", "MCR", "MPR", "ITR"]
    counts = {d: n for d, n in rows}
    return {"domains": [{"domain": d, "count": counts.get(d, 0)} for d in order if d in counts]
            + [{"domain": d, "count": n} for d, n in counts.items() if d not in order],
            "total": sum(counts.values())}


@router.get("/scenes", response_model=list[SceneDto], dependencies=[Depends(get_current_user)])
async def list_scenes(
    domain: str | None = Query(None),
    q: str | None = Query(None),
    session: AsyncSession = Depends(get_session),
):
    """列出场景(可按域 / 关键词过滤)。"""
    stmt = select(StandardScene).where(StandardScene.status == "active")
    if domain:
        stmt = stmt.where(StandardScene.domain == domain)
    if q:
        like = f"%{q}%"
        stmt = stmt.where(StandardScene.name.ilike(like) | StandardScene.code.ilike(like))
    stmt = stmt.order_by(StandardScene.domain, StandardScene.stage, StandardScene.code)
    rows = (await session.execute(stmt)).scalars().all()
    return [_scene_dto(x) for x in rows]


@router.get("/scenes/{scene_id}", response_model=SceneDto, dependencies=[Depends(get_current_user)])
async def get_scene(scene_id: int, session: AsyncSession = Depends(get_session)):
    x = await session.get(StandardScene, scene_id)
    if not x:
        raise HTTPException(404, "场景不存在")
    return _scene_dto(x)


class SceneUpdateBody(BaseModel):
    name: str | None = None
    description: str | None = None
    business_rules: str | None = None
    process: str | None = None
    recommended_fields: list | None = None   # [{name,type,note,required}]
    research_questions: list | None = None   # ["问题1", ...]
    tags: list | None = None                 # ["通用" | "L1/L2/L3/L4"...]
    ai_capabilities: list | None = None      # [ai_capabilities.id ...]


_FIELD_LABELS = {
    "name": "名称", "description": "说明", "business_rules": "业务规则",
    "process": "流程", "recommended_fields": "推荐字段",
    "research_questions": "关键调研问题", "tags": "标签",
    "ai_capabilities": "AI 能力匹配",
}


@router.patch("/scenes/{scene_id}", response_model=SceneDto, dependencies=[Depends(require_admin)])
async def update_scene(
    scene_id: int,
    body: SceneUpdateBody,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """编辑场景内容/标签(仅管理员)。保存写 SceneChange('edit') 留痕并 bump version。"""
    from sqlalchemy.orm.attributes import flag_modified
    x = await session.get(StandardScene, scene_id)
    if not x:
        raise HTTPException(404, "场景不存在")
    changed: list[str] = []
    for field in ("name", "description", "business_rules", "process", "recommended_fields",
                  "research_questions", "tags", "ai_capabilities"):
        val = getattr(body, field)
        if val is not None and getattr(x, field) != val:
            setattr(x, field, val)
            if field in ("recommended_fields", "research_questions", "tags", "ai_capabilities"):
                flag_modified(x, field)
            changed.append(_FIELD_LABELS[field])
    if not changed:
        return _scene_dto(x)
    x.version = (x.version or 1) + 1
    session.add(SceneChange(
        scene_id=x.id, scene_code=x.code, domain=x.domain, change_type="edit",
        summary=f"编辑:{'、'.join(changed)}", created_by=current_user.username,
    ))
    await session.commit()
    await session.refresh(x)
    logger.info("scene_edited", scene_id=scene_id, fields=changed, by=current_user.username)
    return _scene_dto(x)


@router.get("/scenes/{scene_id}/changes", response_model=list[SceneChangeDto],
            dependencies=[Depends(get_current_user)])
async def scene_change_history(scene_id: int, session: AsyncSession = Depends(get_session)):
    rows = (await session.execute(
        select(SceneChange).where(SceneChange.scene_id == scene_id)
        .order_by(SceneChange.created_at.desc())
    )).scalars().all()
    return [SceneChangeDto(
        id=c.id, scene_id=c.scene_id, scene_code=c.scene_code, domain=c.domain,
        change_type=c.change_type, project_name=c.project_name, summary=c.summary,
        created_by=c.created_by, created_at=c.created_at) for c in rows]


@router.get("/scene-changes", response_model=list[SceneChangeDto],
            dependencies=[Depends(get_current_user)])
async def recent_scene_changes(
    limit: int = Query(100, le=500),
    session: AsyncSession = Depends(get_session),
):
    """全库最近变更历史。"""
    rows = (await session.execute(
        select(SceneChange).order_by(SceneChange.created_at.desc()).limit(limit)
    )).scalars().all()
    return [SceneChangeDto(
        id=c.id, scene_id=c.scene_id, scene_code=c.scene_code, domain=c.domain,
        change_type=c.change_type, project_name=c.project_name, summary=c.summary,
        created_by=c.created_by, created_at=c.created_at) for c in rows]


# ── AI 能力目录(场景 AI 能力匹配的可选项底库)────────────────────────────────

class AiCapabilityDto(BaseModel):
    id: int
    domain: str
    agent: str
    skill: str
    status: str
    plan_date: str | None = None
    description: str | None = None
    outputs: list = []


@router.get("/ai-capabilities", response_model=list[AiCapabilityDto],
            dependencies=[Depends(get_current_user)])
async def list_ai_capabilities(session: AsyncSession = Depends(get_session)):
    """纷享已预研 AI 能力目录(场景 AI 能力匹配用)。"""
    rows = (await session.execute(
        select(AiCapability).order_by(AiCapability.sort)
    )).scalars().all()
    return [AiCapabilityDto(
        id=c.id, domain=c.domain, agent=c.agent, skill=c.skill, status=c.status,
        plan_date=c.plan_date, description=c.description, outputs=c.outputs or []) for c in rows]


@router.post("/scenes/ai-match", dependencies=[Depends(require_admin)])
async def ai_match_scenes(
    domain: str | None = Query(None),
    session: AsyncSession = Depends(get_session),
):
    """AI 自动匹配:给场景(可按域)从 AI 能力目录里推荐并落库匹配。仅管理员。"""
    from services.scene_ai_match import auto_match_capabilities
    result = await auto_match_capabilities(session, domain=domain)
    logger.info("scenes_ai_matched", **{k: v for k, v in result.items() if k != "per_domain"})
    return result


# ── 关键调研问题 AI 生成(Part1)──────────────────────────────────────────────

@router.post("/scenes/{scene_id}/gen-questions", dependencies=[Depends(require_admin)])
async def gen_scene_questions(scene_id: int, session: AsyncSession = Depends(get_session)):
    """单场景生成关键调研问题(不落库,前端填入可编辑区)。仅管理员。"""
    from services.scene_questions import gen_questions_for_scene
    x = await session.get(StandardScene, scene_id)
    if not x:
        raise HTTPException(404, "场景不存在")
    questions = await gen_questions_for_scene(session, x)
    return {"questions": questions}


@router.post("/scenes/gen-questions", dependencies=[Depends(require_admin)])
async def batch_gen_scene_questions(
    domain: str | None = Query(None),
    overwrite: bool = Query(False),
    session: AsyncSession = Depends(get_session),
):
    """批量生成关键调研问题并落库(可按域;默认只补空,overwrite 全量重写)。仅管理员。"""
    from services.scene_questions import auto_gen_questions
    result = await auto_gen_questions(session, domain=domain, overwrite=overwrite)
    logger.info("scene_questions_batch", **{k: v for k, v in result.items() if k != "per_domain"})
    return result


# ── 场景导入 / 新增 / 模板下载 ─────────────────────────────────────────────────

_VALID_DOMAINS = {"LTC", "MTL", "MCR", "MPR", "ITR"}

_TEMPLATE_HEADERS = [
    ("domain",     "域(必填)",           "LTC / MTL / MCR / MPR / ITR"),
    ("stage",      "阶段(必填)",         "如 LeadManagement"),
    ("stage_label","阶段显示名",         "如 LeadManagement 线索管理"),
    ("code",       "场景编码(必填)",     "如 LM-01,同域内唯一"),
    ("name",       "场景名称(必填)",     "如 管理线索录入"),
    ("summary",    "阶段定义",           "该阶段的整体说明"),
    ("description","场景说明",           "场景的详细描述"),
    ("business_rules", "业务规则",       "该场景涉及的业务规则"),
    ("process",    "流程",               "该场景的执行流程"),
    ("tags",       "标签",               "多个用分号分隔,如 通用;制造/装备制造"),
]


@router.get("/scenes/import-template", dependencies=[Depends(require_admin)])
async def download_import_template():
    """下载场景导入 Excel 模板(管理员)。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "场景导入"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="D96400", end_color="D96400", fill_type="solid")
    hint_font = Font(color="999999", size=10, italic=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for ci, (_, label, hint) in enumerate(_TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        hint_cell = ws.cell(row=2, column=ci, value=hint)
        hint_cell.font = hint_font
        hint_cell.border = thin_border

    # 示例行
    example = ["LTC", "LeadManagement", "LeadManagement 线索管理", "LM-01",
               "管理线索录入", "线索管理阶段聚焦…", "对来自各渠道的线索进行…",
               "线索必须在24h内分配", "录入→分配→跟进→转化", "通用;制造/装备制造"]
    for ci, val in enumerate(example, 1):
        cell = ws.cell(row=3, column=ci, value=val)
        cell.font = Font(color="888888", size=10)
        cell.border = thin_border

    col_widths = [12, 18, 24, 18, 24, 30, 30, 30, 30, 24]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A3"

    # 说明 sheet
    ws2 = wb.create_sheet("填写说明")
    instructions = [
        "场景导入模板 — 填写说明",
        "",
        "1. 在「场景导入」sheet 中,从第 3 行开始填写(第 2 行为提示,可删除)",
        "2. 必填列:域、阶段、场景编码、场景名称",
        "3. 域只能填:LTC / MTL / MCR / MPR / ITR",
        "4. 场景编码在同域内唯一(如 LM-01),重复编码会更新已有场景",
        "5. 标签列多个标签用英文分号「;」分隔",
        "6. 行业标签格式为层级路径:L1/L2/L3/L4(可只到某一级)",
        "7. 导入会自动跳过空行",
    ]
    for ri, text in enumerate(instructions, 1):
        cell = ws2.cell(row=ri, column=1, value=text)
        if ri == 1:
            cell.font = Font(bold=True, size=13)
        else:
            cell.font = Font(size=11)
    ws2.column_dimensions["A"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "场景导入模板.xlsx"
    encoded = urllib.parse.quote(filename)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


class SceneCreateBody(BaseModel):
    domain: str
    stage: str
    stage_label: str | None = None
    code: str
    name: str
    summary: str | None = None
    description: str | None = None
    business_rules: str | None = None
    process: str | None = None
    tags: list[str] | None = None


@router.post("/scenes", response_model=SceneDto, dependencies=[Depends(require_admin)])
async def create_scene(
    body: SceneCreateBody,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """手动新增单个场景(管理员)。"""
    domain = body.domain.strip().upper()
    code = body.code.strip()
    name = body.name.strip()
    stage = body.stage.strip()

    if not domain or not code or not name or not stage:
        raise HTTPException(400, "域、阶段、编码、名称均为必填")
    if domain not in _VALID_DOMAINS:
        raise HTTPException(400, f"域 '{domain}' 不合法,允许:{', '.join(sorted(_VALID_DOMAINS))}")

    existing = (await session.execute(
        select(StandardScene).where(StandardScene.domain == domain, StandardScene.code == code)
    )).scalar_one_or_none()
    if existing:
        raise HTTPException(409, f"场景编码 {domain}/{code} 已存在")

    scene = StandardScene(
        domain=domain, stage=stage, stage_label=(body.stage_label or "").strip() or None,
        code=code, name=name, summary=(body.summary or "").strip() or None,
        description=(body.description or "").strip() or None,
        business_rules=(body.business_rules or "").strip() or None,
        process=(body.process or "").strip() or None,
        tags=body.tags or [], source_type="standard",
    )
    session.add(scene)
    session.add(SceneChange(
        scene_id=None, scene_code=code, domain=domain, change_type="new",
        summary="后台手动新增", created_by=current_user.username,
    ))
    await session.flush()
    # 回填 scene_id
    change = (await session.execute(
        select(SceneChange).where(SceneChange.scene_code == code, SceneChange.domain == domain)
        .order_by(SceneChange.id.desc()).limit(1)
    )).scalar_one()
    change.scene_id = scene.id
    await session.commit()
    await session.refresh(scene)
    logger.info("scene_created", scene_id=scene.id, code=f"{domain}/{code}", by=current_user.username)
    return _scene_dto(scene)


class ImportResultDto(BaseModel):
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = []


@router.post("/scenes/import", response_model=ImportResultDto, dependencies=[Depends(require_admin)])
async def import_scenes(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """从 Excel 批量导入场景(管理员)。编码重复则更新已有场景。"""
    from openpyxl import load_workbook

    fname = file.filename or ""
    if not fname.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "请上传 .xlsx 文件")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, "文件超过 10MB 限制")

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "无法解析 Excel 文件,请检查格式")

    ws = wb.active
    if not ws:
        raise HTTPException(400, "Excel 无有效工作表")

    field_keys = [h[0] for h in _TEMPLATE_HEADERS]
    result = ImportResultDto()
    rows_processed = 0

    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        cells = {field_keys[ci]: str(row[ci]).strip() if ci < len(row) and row[ci] is not None else ""
                 for ci in range(len(field_keys))}

        domain = cells.get("domain", "").upper()
        code = cells.get("code", "")
        name = cells.get("name", "")
        stage = cells.get("stage", "")

        if not domain and not code and not name:
            continue

        if not domain or not code or not name or not stage:
            result.errors.append(f"第 {ri} 行:域/阶段/编码/名称不能为空")
            result.skipped += 1
            continue
        if domain not in _VALID_DOMAINS:
            result.errors.append(f"第 {ri} 行:域 '{domain}' 不合法")
            result.skipped += 1
            continue

        tags_raw = cells.get("tags", "")
        tags = [t.strip() for t in tags_raw.split(";") if t.strip()] if tags_raw else []

        existing = (await session.execute(
            select(StandardScene).where(StandardScene.domain == domain, StandardScene.code == code)
        )).scalar_one_or_none()

        if existing:
            changed_fields = []
            for field in ("name", "stage", "stage_label", "summary", "description", "business_rules", "process"):
                val = cells.get(field, "")
                if val and val != (getattr(existing, field) or ""):
                    setattr(existing, field, val)
                    changed_fields.append(field)
            if tags and tags != (existing.tags or []):
                existing.tags = tags
                from sqlalchemy.orm.attributes import flag_modified
                flag_modified(existing, "tags")
                changed_fields.append("tags")
            if changed_fields:
                existing.version = (existing.version or 1) + 1
                session.add(SceneChange(
                    scene_id=existing.id, scene_code=code, domain=domain, change_type="edit",
                    summary=f"Excel 导入更新:{'、'.join(changed_fields)}", created_by=current_user.username,
                ))
                result.updated += 1
            else:
                result.skipped += 1
        else:
            scene = StandardScene(
                domain=domain, stage=stage,
                stage_label=cells.get("stage_label") or None,
                code=code, name=name,
                summary=cells.get("summary") or None,
                description=cells.get("description") or None,
                business_rules=cells.get("business_rules") or None,
                process=cells.get("process") or None,
                tags=tags, source_type="standard",
            )
            session.add(scene)
            await session.flush()
            session.add(SceneChange(
                scene_id=scene.id, scene_code=code, domain=domain, change_type="new",
                summary="Excel 导入新增", created_by=current_user.username,
            ))
            result.created += 1

        rows_processed += 1

    if rows_processed == 0 and not result.errors:
        raise HTTPException(400, "未找到有效数据行,请检查模板格式")

    await session.commit()
    wb.close()
    logger.info("scenes_imported", created=result.created, updated=result.updated,
                skipped=result.skipped, errors=len(result.errors), by=current_user.username)
    return result
